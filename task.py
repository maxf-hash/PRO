from openpyxl.workbook import workbook # using openpyxl libary to open excel file
from openpyxl import load_workbook 


wb = load_workbook('employeedata.xlsx') 
ws = wb.active # grab the active worksheet
sheet = wb['Sheet1'] #defining sheet according to the Sheet1 in excel where it is 


for i in range( 2, sheet.max_row+1): #for loop for the different emails and their new values
     cell = sheet.cell(i, 2)
     if 'helpinghands.cm' in cell.value:  #conditional to execute the change in emails
          New = (cell.value).replace('helpinghands.cm', 'handsinhands.org')
          sheet.cell(i, 2).value = New
          
wb.save('updated_emails.xlsx')   # save the file


 